from django.shortcuts import render
from rest_framework.permissions import IsAuthenticated,AllowAny
from rest_framework.views import APIView
# from .utils import scraper_spectra
from rest_framework.response import Response
from .tasks import go_to_sleep,sairtex,webmotors,autoparts,carter,opticat,standard,bwd,wve,oreo,autozone,advance,nepalonline
import io
import csv
import pandas as pd
import openpyxl
from .models import File,Switch_Scrap
from .serializer import FileSer
from .utils2 import check_status
# Create your views here.

class Userdetail(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name
        wb = openpyxl.load_workbook(file)
        worksheet = wb.active
        data = []
        check_status()

        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))

        data = list(filter(None, data))
        data = [d for d in data if d!="None"]
        print(data)
        print("ddddd")

        # a= file.read()
        # print(a)
        # print(type(a))
        # decoded_file = file.read().decode()
        # io_string = io.StringIO(decoded_file)
        # reader = csv.reader(io_string)

        # go_to_sleep.delay(uuid=["16147161387"])
        #
        task = go_to_sleep.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)

class Airtex(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name
        wb = openpyxl.load_workbook(file)
        worksheet = wb.active
        check_status()


        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))

        # go_to_sleep.delay(uuid=["16147161387"])
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]
        task = sairtex.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)


class Mootors(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name
        check_status()

        wb = openpyxl.load_workbook(file)


        worksheet = wb.active

        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))

        # go_to_sleep.delay(uuid=["16147161387"])
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]
        task = webmotors.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)


class Autoparts(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name
        wb = openpyxl.load_workbook(file)
        worksheet = wb.active
        check_status()


        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))

        # go_to_sleep.delay(uuid=["16147161387"])
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]
        task = autoparts.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)


class Carter(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)
        file = request.FILES['file']
        filename = file.name
        check_status()

        wb = openpyxl.load_workbook(file)
        worksheet = wb.active

        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))

        # go_to_sleep.delay(uuid=["16147161387"])
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        task = carter.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)

class Opticat(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name
        check_status()

        wb = openpyxl.load_workbook(file)
        worksheet = wb.active

        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        task = opticat.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)

class Standard(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name
        check_status()

        wb = openpyxl.load_workbook(file)
        worksheet = wb.active

        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))

        # go_to_sleep.delay(uuid=["16147161387"])
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        task = standard.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)

class BWD(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name
        check_status()

        wb = openpyxl.load_workbook(file)
        worksheet = wb.active

        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        # go_to_sleep.delay(uuid=["16147161387"])

        task = bwd.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)

class Alllist(APIView):
    permission_classes = (AllowAny,)
    def get(self, request):
        file = File.objects.all().order_by('-created_date')
        data = FileSer(file,many=True)
        return Response(data.data)

class WVVE(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name

        check_status()

        wb = openpyxl.load_workbook(file)
        worksheet = wb.active

        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        # go_to_sleep.delay(uuid=["16147161387"])

        task = wve.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)



class Oreo(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        check_status()
        filename = file.name


        wb = openpyxl.load_workbook(file)
        worksheet = wb.active

        data = []
        obj,cre = Switch_Scrap.objects.update_or_create(id=1,stop=False)
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        # go_to_sleep.delay(uuid=["16147161387"])

        task = oreo.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)



class Autozone(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name

        check_status()

        wb = openpyxl.load_workbook(file)
        worksheet = wb.active

        data = []
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        # go_to_sleep.delay(uuid=["16147161387"])

        task = autozone.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)



class Advance(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        check_status()
        filename = file.name

        worksheet = wb.active

        data = []
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        # go_to_sleep.delay(uuid=["16147161387"])

        task = advance.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)



class Nepa(APIView):
    permission_classes = (AllowAny,)
    def post(self, request):
        uuid = self.request.query_params.get('id', None)

        file = request.FILES['file']
        filename = file.name

        check_status()

        wb = openpyxl.load_workbook(file,filename)
        worksheet = wb.active

        data = []
        for row in worksheet.iter_rows():
            for cell in row:
                data.append(str(cell.value))
        data = list(filter(None, data))
        data = [d for d in data if d != "None"]

        # go_to_sleep.delay(uuid=["16147161387"])

        task = nepalonline.delay(duration=data,fileName=filename)
        print(task)
        # a = scraper_spectra(uuid)
        content={"task_id":str(task)}
        return Response(content)



class Stop(APIView):
    permission_classes = (AllowAny,)
    def get(self, request):
        a = Switch_Scrap.objects.all()
        if a:
            obj = a[0]
            obj.stop = True
            obj.save()
        else:
            da = Switch_Scrap.objects.create(stop=True)
        return Response("Stop successfully")